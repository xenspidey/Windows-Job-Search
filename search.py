from openpyxl import load_workbook
import glob
import re
import os
from tkinter import *
from tkinter.ttk import Combobox
from tkinter.filedialog import askopenfilenames, askopenfilename, asksaveasfilename
import json
# todo 
# add settings dialog to get location of excel file, also allow for selection of sheet to search within the excel file
key_xlfile = 'xlfile'
value_xlFile = ''
value_xlSheet = ''
searchTerm = ''

def settings():
    '''function for dealing with persistant settings'''

    def read_settings():
        '''read settings from file'''
        with open('config.json', 'w+') as f:
            try:
                config = json.load(f)
            except:
                print('fail_0')
                write_settings('', '')
                return
            try:
                value_xlFile = config['xlfile']
                value_xlSheet = config['xlsheet']
                write_settings('', '')
                return
            except:
                print('fail')

    def write_settings(xlfile, xlsheet):
        '''write config settings to file'''
        config = {'xlfile': xlfile, 'xlsheet': xlsheet}
        with open('config.json', 'w+') as f:
            json.dump(config, f)

    def showSettingsdlg():

        def browse():
            file = askopenfilename(parent=Settingsdlg, filetypes=[('Excel', '.xls, xlsx, xlsm')])
            text_xlFile.delete(0, END)
            text_xlFile.insert(0, file)
            wb = load_workbook(file)
            sheets = wb.get_sheet_names()
            for sheet in sheets:
                combo_sheets.insert(0, sheet)

        def save():
            write_settings(text_xlFile.get(), combo_sheets.get())

        Settingsdlg = Tk()
        Settingsdlg.geometry('300x200')
        text_xlFile = Entry(Settingsdlg)
        button_Browse = Button(Settingsdlg, text='Browse...', command=lambda: browse())
        button_Save = Button(Settingsdlg, text='Save', command=lambda: save())
        combo_sheets = Combobox(Settingsdlg)
        text_xlFile.pack(pady=10, padx=10)
        button_Browse.pack(pady=10, padx=10)
        combo_sheets.pack(pady=10, padx=10)
        button_Save.pack(pady=10, padx=10)
        Settingsdlg.mainloop()


    read_settings()

    if value_xlFile == '' or value_xlSheet == '':
        showSettingsdlg()


def search():
    '''function to search file'''
    def load_searchfile():
        '''load file'''
        wb = load_workbook(value_xlFile)
        sh = wb.get_sheet_by_name(value_xlSheet)
        for row_index in range(sh.get_highest_row()):
            if sh.cell(row=row_index, column=0).value == searchTerm:
                print(row_index)


def main():
    settings()
    root = Tk()
    root.geometry('300x300+1000+500')
    button = Button(root, text='settings', command=lambda: settings())
    button.pack(pady=10, padx=10)
    root.mainloop()

if __name__ == '__main__':
    main()
